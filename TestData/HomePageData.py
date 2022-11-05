import openpyxl
import os

class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        cwd = os.getcwd()
        from pathlib import Path
        path = Path(str(cwd))
        print(str(path.parent)+"\\TestData\\pydemodata.xlsx")
        book = openpyxl.load_workbook(str(cwd)+"\\pydemodata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

